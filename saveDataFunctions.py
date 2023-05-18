def save_data(ws, data: dict, start_row: int, start_column: int):
    from openpyxl import load_workbook, Workbook

    idx = 0
    for key in data.keys():
        # Save the keys to first column of the worksheet starting from the second row
        ws.cell(row=start_row + idx, column=start_column, value=key)

        # Save the values to second column of the worksheet starting from the second row
        ws.cell(row=start_row + idx, column=start_column + 1, value=data[key])

        idx += 1


def save_ages(ws, data: dict, row: int, column: int, title_num: str):
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Data column titles
    ws.cell(row=row, column=column, value="Sock")
    ws.cell(row=row, column=column + 1, value="Age#" + title_num)

    # Save the data
    save_data(ws, data, row + 1, column)

    # Create the table
    tab = Table(displayName="PairsTable" + title_num, ref=f"A{row}"
                                                          f":B{row + len(data)}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style  # Apply the style
    # Add table to the worksheet
    ws.add_table(tab)


def determine_bin_bounds(data: list):
    bin_count = 6
    data_min, data_max = min(data) - 0.5, max(data) + 0.5
    step = (data_max - data_min) / bin_count
    if step < 1:
        step = 1
    else:
        dec = step - int(step)
        step = int(step)
        if dec <= 0.5:
            step += 0.5
        else:
            step += 1

    bins = [data_min + step * i for i in range(bin_count + 1)]

    return bins


def plot_histogram(data: list, title: str = "Title", path: str = "graphs",
                   range_min: int = None, range_max: int = None, show: bool = False, custom_bins=None):
    from matplotlib import pyplot as plt
    from matplotlib.ticker import AutoMinorLocator
    import numpy as np

    if range_min is None:
        range_min = min(data)
    if range_max is None:
        range_max = max(data)

    # plot:
    fig, ax = plt.subplots()

    bin_bounds = determine_bin_bounds(data) if custom_bins is not None else custom_bins
    _, bins, patches = ax.hist(data, bins=bin_bounds,  range=(range_min, range_max), ec='black')

    if len(bins) == 1:
        plt.close()
        return None

    x_ticks = [(bins[idx + 1] + value) / 2 for idx, value in enumerate(bins[:-1])]
    x_ticks_labels = ["[{:.2f}-{:.2f})".format(value, bins[idx+1]) for idx, value in enumerate(bins[:-1])]
    x_ticks_labels[-1] = x_ticks_labels[-1][:-1] + "]"

    plt.xticks(x_ticks, labels=x_ticks_labels, fontsize=6)

    ax.tick_params(axis='x', which='minor', length=0)

    plt.title(title)

    plt.savefig(path + f"/{title}.png")
    if show:
        plt.show()
    plt.close()

    return bin_bounds
