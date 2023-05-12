def save_data(ws, data: dict, start_row: int, start_column: int):
    from openpyxl import load_workbook, Workbook

    idx = 0
    for key in data.keys():
        # Save the keys to first column of the worksheet starting from the second row
        ws.cell(row=start_row + idx, column=start_column, value=key)

        # Save the values to second column of the worksheet starting from the second row
        ws.cell(row=start_row + idx, column=start_column + 1, value=data[key])

        idx += 1


def save_ages(ws, data: dict, row: int, column: int, title_num: str):
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Data column titles
    ws.cell(row=row, column=column, value="Sock")
    ws.cell(row=row, column=column + 1, value="Age#" + title_num)

    # Save the data
    save_data(ws, data, row + 1, column)

    # Create the table
    tab = Table(displayName="PairsTable" + title_num, ref=f"A{row}"
                                                          f":B{row + len(data)}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style  # Apply the style
    # Add table to the worksheet
    ws.add_table(tab)


def save_counts(ws, data: dict, row: int, column: int, title_num: str):
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Data column titles
    ws.cell(row=row, column=column, value="Age#" + title_num)
    ws.cell(row=row, column=column + 1, value="Count")

    # Save the data
    save_data(ws, data, row + 1, column)

    # Create the table
    tab = Table(displayName="PairsTable" + title_num, ref=f"A{row}"
                                                          f":B{row + len(data)}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style  # Apply the style
    # Add table to the worksheet
    ws.add_table(tab)


def bin_algorithm(data: list):
    from math import ceil, sqrt
    bin_count = ceil(sqrt(len(data)))
    data_min, data_max = min(data), max(data)
    step = ceil((data_max - data_min) / bin_count)
    bins = [data_min + (i * step) for i in range(bin_count)]
    bins = list(set(bins))
    bins.sort()
    if bins[-1] < data_max:
        bins[-1] = data_max

    return bins


def plot_histogram(data: list, title: str = "Title", path: str = "graphs",
                   range_min: int = None, range_max: int = None, bin_type="custom", show: bool = False):
    from matplotlib import pyplot as plt
    from matplotlib.ticker import AutoMinorLocator
    import numpy as np

    if range_min is None:
        range_min = min(data)
    if range_max is None:
        range_max = max(data)

    # plot:
    fig, ax = plt.subplots()

    if bin_type == "custom":
        bin_bounds = bin_algorithm(data)
    else:
        bin_bounds = bin_type

    _, bins, patches = ax.hist(data, bins=bin_bounds,  range=(range_min, range_max), ec='black')
    # x ticks
    x_ticks = [(bins[idx+1] + value)/2 for idx, value in enumerate(bins[:-1])]
    x_ticks_labels = ["[{:.0f}-{:.0f})".format(value, bins[idx+1]) for idx, value in enumerate(bins[:-1])]
    x_ticks_labels[-1] = x_ticks_labels[-1][:-1] + "]"
    if bin_type == "custom":
        plt.xticks(x_ticks, labels=x_ticks_labels, fontsize=8)
    else:
        plt.xticks(x_ticks, labels=x_ticks_labels, fontsize=4)

    ax.tick_params(axis='x', which='minor', length=0)

    plt.title(title)

    plt.savefig(path + f"/{title}.png")
    if show:
        plt.show()
    plt.close()
